from pathlib import Path
from typing import List, Dict, Any, Iterable
import pandas as pd
import string


# =========================
# Утилиты для CSV
# =========================
def _smart_read_csv(p: Path) -> pd.DataFrame:
    """Чтение CSV с попытками разных кодировок и разделителей."""
    for enc in ("utf-8-sig", "cp1251", "utf-8"):
        for sep in (None, ";", ","):
            try:
                return pd.read_csv(p, sep=sep, engine="python", encoding=enc)
            except Exception:
                continue
    return pd.read_csv(p)  # даст явную ошибку


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Нормализуем заголовки: нижний регистр, обрезка пробелов, ё→е."""
    cols = []
    for c in df.columns:
        norm = str(c).strip().lower().replace("ё", "е")
        cols.append(norm)
    out = df.copy()
    out.columns = cols
    return out


def _pick_column(df_norm: pd.DataFrame, candidates: Iterable[str]) -> str:
    """Ищем колонку: точное совпадение → подстрока → ошибка."""
    for cand in candidates:
        if cand in df_norm.columns:
            return cand
    for col in df_norm.columns:
        for cand in candidates:
            if cand in col:
                return col
    raise KeyError(f"Не найдена колонка из списка: {list(candidates)}. Доступные: {list(df_norm.columns)}")


def _count_in_df(df: pd.DataFrame) -> Dict[str, int]:
    """Подсчёт: выполнено всего, постоматы (П), ПВЗ, доставки и заявки."""
    df_norm = _normalize_columns(df)
    status_col = _pick_column(df_norm, ["статус задания", "статус", "статус_задания"])
    type_col   = _pick_column(df_norm, ["тип адреса", "тип_адреса", "тип точки", "тип_точки", "тип"])

    # фильтруем выполненные
    status_series = df_norm[status_col].astype(str).str.strip().str.lower().str.replace("ё", "е", regex=False)
    done_mask = status_series.eq("выполнено")

    type_series = df_norm[type_col].astype(str).str.strip().str.upper()

    # категории
    postomats = int((done_mask & type_series.eq("П")).sum())
    pvz = int((done_mask & type_series.eq("ПВЗ")).sum())
    deliveries = int((done_mask & type_series.eq("Д")).sum())
    orders = int((done_mask & type_series.eq("З")).sum())

    # остальные (всё выполненное, что не попало в эти категории)
    total_completed = int(done_mask.sum())
    categorized = postomats + pvz + deliveries + orders


    return {
        "total_completed": total_completed,
        "postomats": postomats,
        "pvz": pvz,
        "deliveries": deliveries,
        "orders": orders,
    }


def analyze_csvs(paths: List[Path]) -> Dict[str, Any]:
    """Считает итоги по CSV среди выбранных путей (Excel игнорируется)."""
    totals = {"total_completed": 0, "postomats": 0, "pvz": 0, "deliveries": 0, "orders": 0}
    per_file, errors = [], []

    for p in paths:
        if p.suffix.lower() != ".csv":
            continue
        try:
            df = _smart_read_csv(p)
            counts = _count_in_df(df)
            per_file.append({"file": str(p), **counts})
            for k in totals:
                totals[k] += counts[k]
        except Exception as e:
            errors.append({"file": str(p), "error": str(e)})

    return {"totals": totals, "per_file": per_file, "errors": errors}


# =========================
# Остальная логика (КТ и ПМ) — без изменений
# =========================

KEEP_COLUMNS_KT = [
    "Номер заказа",
    "Код офиса местонахождения",
    "Контрольная точка",
    "Дней в КТ",
    "Тип заказа",
]


def _autosize_columns_xlsx(xlsx_path: Path, df_out: pd.DataFrame) -> None:
    import openpyxl
    from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    for col_idx, col_name in enumerate(df_out.columns, start=1):
        max_len = len(str(col_name)) if col_name is not None else 0
        for val in df_out[col_name].astype(str).fillna("").values:
            ln = len(str(val))
            if ln > max_len:
                max_len = ln
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
    wb.save(xlsx_path)


def process_kt_excels(paths: List[Path], keep_columns: List[str] = None) -> Dict[str, Any]:
    if keep_columns is None:
        keep_columns = KEEP_COLUMNS_KT
    saved, skipped, errors = [], [], []
    for p in paths:
        ext = p.suffix.lower()
        if ext not in (".xlsx", ".xls", ".xlsm"):
            skipped.append({"file": str(p), "reason": "Не Excel"})
            continue
        try:
            import openpyxl
            df = pd.read_excel(p)
            existing = [c for c in keep_columns if c in df.columns]
            if not existing:
                raise KeyError(f"Нет нужных столбцов. Есть: {list(df.columns)}; Ожидались: {keep_columns}")
            df_out = df[existing].copy()
            out_path = p.with_name(f"{p.stem}_KT.xlsx")
            df_out.to_excel(out_path, index=False)
            _autosize_columns_xlsx(out_path, df_out)
            saved.append({"file": str(p), "saved_as": str(out_path), "kept_columns": existing})
        except Exception as e:
            errors.append({"file": str(p), "error": str(e)})
    return {"saved": saved, "skipped": skipped, "errors": errors}


# =========================
# ПМ (последняя миля)
# =========================
PM_CODES = {
    "Декабрьская": "MSK650",
    "Живова": "MSK963",
    "Мневники": "MSK1125",
    "Твардовского": "MSK2469",
}
PM_METRIC_RAW = "Ср. срок на последней миле для 2 якоря без СДД, дн"


def _norm_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower().replace("ё", "е")
    table = str.maketrans("", "", string.punctuation + " ")
    return s.translate(table)


def _find_metric_col_loose(df: pd.DataFrame) -> str:
    target = _norm_text(PM_METRIC_RAW)
    for c in df.columns:
        if _norm_text(c) == target:
            return c
    for c in df.columns:
        if (nc := _norm_text(c)) and (target in nc or nc in target):
            return c
    keywords = ["последней", "мил", "якор", "сдд", "срок"]
    for c in df.columns:
        nc = _norm_text(c)
        if sum(kw in nc for kw in keywords) >= 3:
            return c
    raise KeyError(f"Не найдена колонка метрики: «{PM_METRIC_RAW}». Колонки: {list(df.columns)}")


def _find_code_col_loose(df: pd.DataFrame, codes: List[str]) -> str:
    best_col, best_hits = None, -1
    upcodes = [c.strip().upper() for c in codes]
    for col in df.columns:
        ser = df[col].astype(str).str.strip().str.upper()
        hits = sum(ser.eq(code).any() for code in upcodes)
        if hits > best_hits:
            best_hits, best_col = hits, col
    if best_col is None or best_hits <= 0:
        for col in df.columns:
            if df[col].astype(str).str.contains(r"\bMSK\d+\b", regex=True, na=False).any():
                return col
        raise KeyError("Не найдена колонка с кодами MSK***.")
    return best_col


def _extract_pm_from_df(df: pd.DataFrame) -> Dict[str, Any]:
    metric_col = _find_metric_col_loose(df)
    code_col = _find_code_col_loose(df, list(PM_CODES.values()))
    values: Dict[str, Any] = {}
    ser_code = df[code_col].astype(str).str.strip().str.upper()
    for name, code in PM_CODES.items():
        row = df.loc[ser_code == code]
        values[name] = None if row.empty else row.iloc[0][metric_col]
    return values


def analyze_pm_excels(paths: List[Path]) -> Dict[str, Any]:
    results, errors, skipped = [], [], []
    for p in paths:
        ext = p.suffix.lower()
        if ext not in (".xlsx", ".xls", ".xlsm"):
            skipped.append({"file": str(p), "reason": "Не Excel"})
            continue
        try:
            import openpyxl
            xls = pd.read_excel(p, sheet_name=None)
            best_sheet, best_vals, best_score = None, None, -1
            for sheet_name, df in xls.items():
                try:
                    vals = _extract_pm_from_df(df)
                    score = sum(v is not None for v in vals.values())
                    if score > best_score:
                        best_sheet, best_vals, best_score = sheet_name, vals, score
                        if score == len(PM_CODES):
                            break
                except Exception:
                    continue
            if best_vals is None:
                raise KeyError("Не удалось извлечь ПМ-значения ни с одного листа.")
            results.append({"file": str(p), "sheet": best_sheet, "values": best_vals})
        except Exception as e:
            errors.append({"file": str(p), "error": str(e)})
    return {"results": results, "errors": errors, "skipped": skipped}